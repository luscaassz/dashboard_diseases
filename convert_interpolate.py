#!/usr/bin/env python3
"""
Normalize Excel and CSV files to monthly long format.

For each input file (xlsx or csv) this script:
 - If .xlsx: tries to read the best sheet (prefers 'MEDIA'), saves a raw CSV copy to output_dir/raw/<stem>.csv
 - If .csv: reads it directly (and copies to output_dir/raw/<stem>.csv for consistency)
 - Detects id columns (cod_mun, nome_mun) and date-like columns (annual or monthly)
 - Converts wide->long, parses dates, converts values to numeric
 - Builds monthly series (freq='MS') for each municipality, interpolates by time, ffill/bfill edges
 - Saves final long CSV: output_dir/<stem>_monthly_long.csv with columns:
       cod_mun, nome_mun, date (YYYY-MM-01), value

Usage:
    python normalize_to_monthly_long.py --input_dir data --output_dir data/monthly --dens_offset 2000
"""
import argparse
from pathlib import Path
import re
import pandas as pd
import numpy as np
import warnings
import sys

pd.options.mode.chained_assignment = None

# ------------------ Helpers (kept & improved from your version) ------------------
def parse_header_to_date(h: object, dens_offset: int = 2000):
    """Tenta converter cabeçalhos como '01-1999.', '02.1999', '1999-01', 'DS_POP_00' em pd.Timestamp."""
    if h is None:
        return None
    if isinstance(h, pd.Timestamp):
        return pd.Timestamp(year=h.year, month=h.month, day=1)

    s = str(h).strip()
    if s == "":
        return None

    # limpar
    s = s.strip().rstrip(".").replace("..", ".").strip()

    # 01-1999, 1-1999, 01/1999, 01.1999
    m = re.match(r"^\s*0?([1-9]|1[0-2])\s*[-/\.]\s*(\d{4})\s*$", s)
    if m:
        return pd.Timestamp(year=int(m.group(2)), month=int(m.group(1)), day=1)

    # 1999-01, 1999/01, 1999.01
    m2 = re.match(r"^\s*(\d{4})\s*[-/\.]\s*0?([1-9]|1[0-2])\s*$", s)
    if m2:
        return pd.Timestamp(year=int(m2.group(1)), month=int(m2.group(2)), day=1)

    # Ano puro (1999)
    if re.match(r"^\s*\d{4}\s*$", s):
        return pd.Timestamp(year=int(s), month=1, day=1)

    # DS_POP_00 / DS_POP_23 -> offset
    if re.search(r'ds[_\s-]*pop', s, flags=re.I):
        m = re.search(r'(\d{1,2})\s*$', s)
        if m:
            ano = dens_offset + int(m.group(1))
            return pd.Timestamp(year=ano, month=1, day=1)

    # Ano/MonAbbrev (2000/Jan ou 2000/Jan.)
    m3 = re.match(r'^\s*(\d{4})\s*/\s*([A-Za-zÀ-ÿ]+)\.?\s*$', s)
    if m3:
        year = int(m3.group(1))
        mon_abbr = m3.group(2).strip().upper()[:3]
        meses_pt = {'JAN':1,'FEV':2,'MAR':3,'ABR':4,'MAI':5,'JUN':6,'JUL':7,'AGO':8,'SET':9,'OUT':10,'NOV':11,'DEZ':12}
        meses_en = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
        mon = meses_pt.get(mon_abbr) or meses_en.get(mon_abbr)
        if mon:
            return pd.Timestamp(year=year, month=mon, day=1)

    # tentativa genérica com pandas
    try:
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if not pd.isna(dt):
            return pd.Timestamp(year=dt.year, month=dt.month if dt.month>0 else 1, day=1)
    except Exception:
        pass

    return None


def to_numeric_robust(v):
    """Converte string numérica com vírgula / milhares para float; caso contrário NaN."""
    if v is None:
        return np.nan
    if isinstance(v, (int, float, np.number)) and not pd.isnull(v):
        return float(v)
    s = str(v).strip()
    if s == '':
        return np.nan
    # handle thousand separators pt-br 1.234,56
    if re.match(r'^\d{1,3}(\.\d{3})+,\d+$', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.').replace(' ', '')
    # remove non-number trailing/leading
    try:
        return float(s)
    except:
        m = re.search(r'[-+]?\d*\.\d+|\d+', s)
        return float(m.group(0)) if m else np.nan


def detect_id_columns(cols):
    """Heurística para encontrar coluna de código e nome do município."""
    code_col = None
    name_col = None
    for c in cols:
        if re.search(r'cd[_\s-]*mun|^cod[_\s-]*mun|^codmun|^cdmun|^codigo|^cod', str(c), flags=re.I):
            code_col = c
            break
    for c in cols:
        if re.search(r'\bmun\b|municipio|nome|nm[_\s-]*mun|nm_mun|nm_munic', str(c), flags=re.I):
            name_col = c
            break
    if code_col is None:
        code_col = cols[0] if cols else None
    return code_col, name_col


# ------------------ Core normalization logic (supports XLSX and CSV) ------------------
def read_xlsx_best(path: Path):
    """Tentativa robusta de leitura de .xlsx: prefer 'MEDIA', fallback heurísticas."""
    import openpyxl
    # tentativas com pandas direto (sheet MEDIA)
    df = None
    chosen = None
    try:
        try:
            df = pd.read_excel(path, sheet_name='MEDIA', engine='openpyxl', header=0, dtype=object)
            chosen = 'MEDIA'
            return df, chosen
        except Exception:
            pass

        # listar sheets via pandas.ExcelFile (openpyxl engine)
        try:
            xls = pd.ExcelFile(path, engine='openpyxl')
            # procura nome 'MEDIA' (case-insensitive)
            cand = [n for n in xls.sheet_names if str(n).strip().upper() == 'MEDIA']
            if cand:
                df = pd.read_excel(path, sheet_name=cand[0], engine='openpyxl', header=0, dtype=object)
                chosen = cand[0]
                return df, chosen
        except Exception:
            pass

        # fallback com openpyxl.load_workbook
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        # prefer sheet with many rows/cols
        for name in sheetnames:
            try:
                ws = wb[name]
                max_row = getattr(ws, 'max_row', 0)
                max_col = getattr(ws, 'max_column', 0)
                if max_row >= 2 and max_col >= 2:
                    # tentar ler com pandas
                    try:
                        df_try = pd.read_excel(path, sheet_name=name, engine='openpyxl', header=0, dtype=object)
                        if df_try is not None and df_try.shape[1] > 0:
                            df = df_try
                            chosen = name
                            return df, chosen
                    except Exception:
                        continue
            except Exception:
                continue

        # final fallback: tentar qualquer sheet
        for name in sheetnames:
            try:
                df_try = pd.read_excel(path, sheet_name=name, engine='openpyxl', header=0, dtype=object)
                if df_try is not None and df_try.shape[1] > 0:
                    df = df_try
                    chosen = name
                    return df, chosen
            except Exception:
                continue

    except Exception as e:
        print("  read_xlsx_best error:", e)
    return None, None


def normalize_dataframe(df: pd.DataFrame, stem: str, output_dir: Path, dens_offset: int = 2000, code_col_override=None, name_col_override=None):
    """Converte um DataFrame (wide) para monthly long format e salva o arquivo final."""
    if df is None or df.shape[0] == 0 or df.shape[1] == 0:
        print("  DataFrame vazio; pulando.")
        return False

    # normaliza colnames
    cols = [str(c).strip() for c in df.columns]
    df.columns = cols

    # detect id columns
    code_col, name_col = detect_id_columns(cols)
    if code_col_override:
        code_col = code_col_override
    if name_col_override:
        name_col = name_col_override
    print(f"  detect -> code_col = {code_col}, name_col = {name_col}")

    # detect date-like columns
    date_cols = []
    parsed_map = {}
    for c in cols:
        if c == code_col or (name_col and c == name_col):
            continue
        parsed = parse_header_to_date(c, dens_offset=dens_offset)
        if parsed is not None:
            date_cols.append(c)
            parsed_map[c] = parsed

    # fallback: columns containing digits and not id columns
    if not date_cols:
        date_cols = [c for c in cols if re.search(r'\d', c) and c not in [code_col, name_col]]
        # attempt to parse them now
        parsed_map = {c: parse_header_to_date(c, dens_offset=dens_offset) for c in date_cols}
        date_cols = [c for c in date_cols if parsed_map.get(c) is not None]

    if not date_cols:
        print("  Nenhuma coluna tipo-data detectada; pulando normalização deste arquivo.")
        return False

    print(f"  Encontradas {len(date_cols)} colunas de data (ex: {date_cols[:6]})")

    # wide->long melt
    id_vars = [code_col] + ([name_col] if name_col and name_col != code_col else [])
    # Se code_col for None, usar primeira coluna
    id_vars = [v for v in id_vars if v is not None]
    if not id_vars:
        id_vars = [cols[0]]
    long = df.melt(id_vars=id_vars, value_vars=date_cols, var_name='raw_date', value_name='raw_value')

    # parse date and numeric conversion
    long['parsed_date'] = long['raw_date'].apply(lambda x: parsed_map.get(x) if x in parsed_map else parse_header_to_date(x, dens_offset=dens_offset))
    long['value'] = long['raw_value'].apply(to_numeric_robust)

    # drop rows without parsed_date
    before = len(long)
    long = long.dropna(subset=['parsed_date']).copy()
    dropped = before - len(long)
    if dropped:
        print(f"  Dropped {dropped} rows sem data parseável.")

    # group by municipality code
    results = []
    grouped = long.groupby(id_vars[0])
    for code, group in grouped:
        # average if same date duplicates
        ser = group.groupby('parsed_date')['value'].mean().sort_index()
        if ser.dropna().empty:
            continue

        start = ser.index.min()
        end = ser.index.max()
        monthly_index = pd.date_range(start=start, end=end, freq='MS')
        ser_full = ser.reindex(ser.index.union(monthly_index)).sort_index()
        # interpolate by time; then ensure monthly_index present
        ser_monthly = ser_full.interpolate(method='time').reindex(monthly_index)
        ser_monthly = ser_monthly.ffill().bfill()

        # try get nome_mun
        nome = ''
        if len(id_vars) > 1 and id_vars[1] in group.columns:
            try:
                non_empty_names = group[id_vars[1]].astype(str).replace('nan', '').replace('None', '').str.strip()
                non_empty_names = non_empty_names[non_empty_names != '']
                if not non_empty_names.empty:
                    nome = str(non_empty_names.iloc[0])
            except Exception:
                nome = ''

        # build df
        df_out = pd.DataFrame({
            'cod_mun': str(code),
            'nome_mun': nome,
            'date': ser_monthly.index,
            'value': ser_monthly.values
        })
        results.append(df_out)

    if not results:
        print("  Nenhuma série válida gerada.")
        return False

    final = pd.concat(results, ignore_index=True)
    final['date'] = pd.to_datetime(final['date']).dt.to_period('M').dt.to_timestamp()
    # normalize cod_mun string cleanup (remover .0 de floats)
    final['cod_mun'] = final['cod_mun'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

    # salvar resultado final
    output_dir.mkdir(parents=True, exist_ok=True)
    out_file = output_dir / f"{stem}_monthly_long.csv"
    final.to_csv(out_file, index=False, date_format='%Y-%m-%d')
    print(f"  Saved normalized monthly-long CSV: {out_file} (rows: {len(final)})")
    return True


def process_file(path: Path, output_dir: Path, raw_dir: Path, dens_offset: int = 2000):
    """Processa um arquivo (xlsx ou csv). Retorna True se sucesso."""
    stem = path.stem
    print(f"\nProcessing: {path.name}")
    try:
        if path.suffix.lower() in ['.xls', '.xlsx', '.xlsm', '.xlsb']:
            # Ler xlsx robustamente e também salvar CSV copia raw
            df, sheet = read_xlsx_best(path)
            if df is None:
                # última tentativa: usar pandas.read_excel sem engine fallback
                try:
                    df = pd.read_excel(path, header=0, dtype=object)
                except Exception as e:
                    print(f"  ERROR: Não foi possível ler {path.name} como Excel: {e}")
                    return False

            # Save a raw CSV copy for record (normalize columns then write)
            try:
                raw_dir.mkdir(parents=True, exist_ok=True)
                raw_csv = raw_dir / f"{stem}.csv"
                # ensure columns strings
                df_raw = df.copy()
                df_raw.columns = [str(c).strip() for c in df_raw.columns]
                df_raw.to_csv(raw_csv, index=False, encoding='utf-8')
                print(f"  Saved raw CSV copy: {raw_csv}")
            except Exception as e:
                print(f"  Warning: não consegui salvar raw CSV copy: {e}")

            # Now normalize
            ok = normalize_dataframe(df, stem, output_dir, dens_offset=dens_offset)
            return ok

        elif path.suffix.lower() == '.csv':
            # read csv robustly (attempt to autodetect delimiter, header presence)
            # try common encodings/delimiters
            read_success = False
            df = None
            encodings = ['utf-8', 'latin1', 'cp1252']
            delims = [',',';','\t','|']
            for enc in encodings:
                for d in delims:
                    try:
                        df_try = pd.read_csv(path, dtype=object, encoding=enc, sep=d, low_memory=False)
                        # heuristic: require at least 2 columns and some rows
                        if df_try.shape[1] >= 2 and df_try.shape[0] >= 1:
                            df = df_try
                            read_success = True
                            used_enc = enc
                            used_delim = d
                            break
                    except Exception:
                        continue
                if read_success:
                    break
            if not read_success:
                # final fallback: pandas autodetect
                try:
                    df = pd.read_csv(path, dtype=object, low_memory=False)
                    used_enc = 'fallback'
                    used_delim = ','
                    read_success = True
                except Exception as e:
                    print(f"  ERROR reading CSV {path.name}: {e}")
                    return False

            print(f"  CSV read with encoding={used_enc} delim='{used_delim}' shape={df.shape}")

            # Save a raw copy to raw_dir (normalize colnames) for consistency
            try:
                raw_dir.mkdir(parents=True, exist_ok=True)
                raw_csv = raw_dir / f"{stem}.csv"
                df_copy = df.copy()
                df_copy.columns = [str(c).strip() for c in df_copy.columns]
                df_copy.to_csv(raw_csv, index=False, encoding='utf-8')
                print(f"  Saved raw CSV copy: {raw_csv}")
            except Exception as e:
                print(f"  Warning: could not save raw copy: {e}")

            ok = normalize_dataframe(df, stem, output_dir, dens_offset=dens_offset)
            return ok
        else:
            print(f"  Unsupported file extension: {path.suffix}. Skipping.")
            return False
    except Exception as e:
        print(f"  ERROR processing file {path.name}: {e}")
        return False


# ------------------ CLI / main ------------------
def main():
    parser = argparse.ArgumentParser(description="Normalize Excel and CSV files to monthly long CSVs")
    parser.add_argument("--input_dir", type=str, default="data", help="Input folder with .xlsx/.csv files")
    parser.add_argument("--output_dir", type=str, default="data/monthly", help="Output folder for monthly long CSVs")
    parser.add_argument("--raw_subdir", type=str, default="raw", help="Subfolder inside output_dir to save raw CSV copies of inputs")
    parser.add_argument("--pattern", type=str, default="*", help="Glob pattern for files (e.g. *.xlsx or * to include csv/xlsx)")
    parser.add_argument("--dens_offset", type=int, default=2000, help="Offset for DS_POP_XX -> year")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    raw_dir = output_dir / args.raw_subdir

    files = sorted([p for p in input_dir.glob(args.pattern) if p.suffix.lower() in ['.xlsx','.xls','.xlsm','.xlsb','.csv']])
    if not files:
        print("No files found in", input_dir, "with pattern", args.pattern)
        return

    print(f"Found {len(files)} files to process.")
    success = 0
    for f in files:
        ok = process_file(f, output_dir, raw_dir, dens_offset=args.dens_offset)
        if ok:
            success += 1

    print(f"\nDone. Successfully processed {success}/{len(files)} files.")


if __name__ == "__main__":
    main()
